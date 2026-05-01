import pandas as pd
import sqlite3
import os
import json
import re
from datetime import datetime
import sys

project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(project_root)
from src.llm.goals_predictor import GoalsPredictor

def predict_goals_for_date(target_date: str, is_repredict: bool = False):
    print(f"开始{'重新' if is_repredict else ''}预测 {target_date} 的进球数...")
    
    xl_path = os.path.join(project_root, 'docs', 'foot_prediction.xlsx')
    if not os.path.exists(xl_path):
        return False, "找不到 foot_prediction.xlsx 文件"
        
    try:
        from openpyxl import load_workbook
        wb = load_workbook(xl_path)
        
        # 尝试通过表名匹配日期，或者遍历所有表找到包含该日期的行
        target_ws = None
        
        # 1. 优先尝试直接用日期作为表名查找
        for sheet_name in wb.sheetnames:
            if target_date in sheet_name:
                target_ws = wb[sheet_name]
                break
                
        # 如果没有找到同名 sheet，我们在后面逻辑里去遍历所有 sheet，这里先不报错
    except Exception as e:
        return False, f"加载 Excel 失败: {e}"

    target_rows = []
    headers_info = {} # 记录每个 sheet 的表头信息
    
    # 如果没找到特定的 sheet，就遍历所有的 sheet (排除隐藏或说明性质的 sheet，一般以日期命名)
    sheets_to_check = [target_ws] if target_ws else [wb[name] for name in wb.sheetnames]
    
    print(f"DEBUG: 开始检查 {len(sheets_to_check)} 个 sheet")
    for ws in sheets_to_check:
        print(f"DEBUG: 正在检查 sheet: {ws.title}")
        # 找到列索引
        headers = {str(cell.value).strip(): idx for idx, cell in enumerate(ws[1], 1) if cell.value}
        print(f"DEBUG: [{ws.title}] 找到的表头: {list(headers.keys())}")
        
        date_col = headers.get('日期')
        code_col = headers.get('编码')
        home_col = headers.get('主队')
        away_col = headers.get('客队')
        pan_col = headers.get('进球盘口')
        diff_col = headers.get('预测差异百分比')
        trend_col = headers.get('倾向')
        # 处理倾向列可能存在的空格
        if not trend_col and ' 倾向' in headers:
            trend_col = headers.get(' 倾向')
            
        pred_col_name = '重新预测' if is_repredict else '预测进球数'
        fund_col_name = '基本面预测进球数'
        
        pred_col = headers.get(pred_col_name)
        fund_col = headers.get(fund_col_name)
        
        # 兼容性处理：如果没找到 pred_col，可能列名有轻微差异（比如空格）
        if not pred_col:
            for k, v in headers.items():
                if pred_col_name in str(k).strip():
                    pred_col = v
                    break
                    
        if not fund_col:
            for k, v in headers.items():
                if fund_col_name in str(k).strip():
                    fund_col = v
                    break
                    
        # 再一次兼容性：尝试遍历前 30 列的第 1 行，直接找名字包含它的列
        if not pred_col or not fund_col:
            for col_idx in range(1, 31):
                cell_val = ws.cell(row=1, column=col_idx).value
                if cell_val:
                    clean_val = str(cell_val).strip()
                    if not pred_col and pred_col_name in clean_val:
                        pred_col = col_idx
                    if not fund_col and fund_col_name in clean_val:
                        fund_col = col_idx
                        
        # 如果依然没有 fund_col，我们需要在最后增加一列
        if not fund_col:
            fund_col = ws.max_column + 1
            ws.cell(row=1, column=fund_col).value = fund_col_name
        
        if not all([date_col, code_col, pred_col]):
            print(f"DEBUG: [{ws.title}] 跳过：未找到必需列。找到 date={date_col}, code={code_col}, pred={pred_col}")
            continue # 这个 sheet 格式不对，跳过
            
        headers_info[ws.title] = {
            'date_col': date_col, 'code_col': code_col, 'pan_col': pan_col,
            'diff_col': diff_col, 'trend_col': trend_col, 'pred_col': pred_col, 'fund_col': fund_col
        }

        # 获取目标日期的比赛行
        match_count_in_sheet = 0
        for row_idx in range(2, ws.max_row + 1):
            cell_date = ws.cell(row=row_idx, column=date_col).value
            if not cell_date:
                continue
            
            # 统一格式化日期为 YYYY-MM-DD
            if isinstance(cell_date, datetime):
                row_date = cell_date.strftime('%Y-%m-%d')
            else:
                try:
                    row_date = pd.to_datetime(str(cell_date)).strftime('%Y-%m-%d')
                except:
                    row_date = str(cell_date).strip()[:10]
            
            # 增加对单元格日期的调试打印
            if row_idx < 5:
                print(f"DEBUG: [{ws.title}] Row {row_idx} date: {row_date} (Original: {type(cell_date)})")
                    
            if row_date == target_date:
                match_count_in_sheet += 1
                code = str(ws.cell(row=row_idx, column=code_col).value).strip()
                pan = ws.cell(row=row_idx, column=pan_col).value if pan_col else ''
                diff = ws.cell(row=row_idx, column=diff_col).value if diff_col else ''
                trend = ws.cell(row=row_idx, column=trend_col).value if trend_col else ''
                
                target_rows.append({
                    'ws': ws, # 记录属于哪个 sheet
                    'pred_col': pred_col, # 记录该 sheet 对应的预测列
                    'fund_col': fund_col,
                    'row_idx': row_idx,
                    'match_num': code,
                    'pan': str(pan).strip() if pd.notna(pan) and str(pan).strip() != 'nan' else '',
                    'diff': str(diff).strip() if pd.notna(diff) and str(diff).strip() != 'nan' else '',
                    'trend': str(trend).strip() if pd.notna(trend) and str(trend).strip() != 'nan' else ''
                })
        print(f"DEBUG: [{ws.title}] 找到符合日期 {target_date} 的比赛共 {match_count_in_sheet} 场")

    if not target_rows:
        return False, f"Excel 中没有找到日期为 {target_date} 的比赛记录"

    # 读取数据库
    db_path = os.path.join(project_root, 'data', 'football.db')
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # 获取比赛的 raw_data
    goals_predictor = GoalsPredictor()
    success_count = 0
    
    for row_data in target_rows:
        match_num = row_data['match_num']
        # 查询数据库，找到最接近这一天的预测记录的 raw_data
        cursor.execute("""
            SELECT raw_data FROM match_predictions 
            WHERE match_num = ? AND (date(match_time) = ? OR date(created_at) = ?)
            ORDER BY created_at DESC LIMIT 1
        """, (match_num, target_date, target_date))
        
        db_res = cursor.fetchone()
        if not db_res:
            # 宽泛查询
            cursor.execute("SELECT raw_data FROM match_predictions WHERE match_num = ? ORDER BY created_at DESC LIMIT 1", (match_num,))
            db_res = cursor.fetchone()
            
        if db_res and db_res[0]:
            try:
                raw_data = json.loads(db_res[0])
                # 构造 match_data
                match_data = raw_data.copy()
                match_data['goals_pan'] = row_data['pan']
                match_data['goals_diff_percent'] = row_data['diff']
                match_data['goals_trend'] = row_data['trend']
                
                # 如果没有进球盘口等数据，跳过
                if not match_data.get('goals_pan') and not match_data.get('goals_diff_percent') and not match_data.get('goals_trend'):
                    print(f"比赛 {match_num} 缺少机构盘口数据，跳过")
                    continue
                    
                goals_pred, _ = goals_predictor.predict(match_data)
                
                # 解析预测结果
                try:
                    res_dict = json.loads(goals_pred)
                    stat_goals = res_dict.get('statistical_goals')
                    fund_report = res_dict.get('fundamental_report', '')
                except Exception as e:
                    stat_goals = None
                    fund_report = goals_pred

                clean_pred = re.sub(r'\*+', '', fund_report).strip()
                match_goals = re.search(r'【进球数预测】\s*([^\n]+)', clean_pred)
                extracted_goals = None
                if match_goals:
                    extracted_goals = match_goals.group(1).strip()
                else:
                    match_goals = re.search(r'进球数预测.*?\n([^\n]+)', clean_pred)
                    if match_goals:
                        extracted_goals = match_goals.group(1).strip()
                        
                # 如果没有匹配到统计结果，或者解析失败，降级处理
                if stat_goals:
                    row_data['ws'].cell(row=row_data['row_idx'], column=row_data['pred_col']).value = stat_goals
                elif extracted_goals:
                    row_data['ws'].cell(row=row_data['row_idx'], column=row_data['pred_col']).value = extracted_goals
                
                if extracted_goals:
                    row_data['ws'].cell(row=row_data['row_idx'], column=row_data['fund_col']).value = extracted_goals
                
                success_count += 1
                
                # 尝试从 fund_report 提取统计数据洞察以打印到控制台
                insight_text = ""
                insight_match = re.search(r'【统计数据洞察】\s*(.*?)(?=\n\n|\n\*\*|$)', fund_report, re.DOTALL)
                if insight_match:
                    insight_text = insight_match.group(1).strip().replace('\n', ' | ')
                    
                print(f"比赛 {match_num} 统计预测: {stat_goals} | 基本面预测: {extracted_goals}")
                if insight_text:
                    print(f"  └─ 洞察: {insight_text}")
                    
            except Exception as e:
                print(f"处理比赛 {match_num} 时出错: {e}")
                continue
        else:
            print(f"数据库中未找到比赛 {match_num} 的记录")

    conn.close()
    
    if success_count > 0:
        wb.save(xl_path)
        return True, f"成功{'重新' if is_repredict else ''}预测并回写了 {success_count} 场比赛的进球数"
    else:
        return False, "没有成功预测任何比赛，可能是因为缺少机构进球数盘口数据或数据库中未找到对应比赛。"

if __name__ == '__main__':
    # 简单的测试入口
    if len(sys.argv) > 1:
        date = sys.argv[1]
        is_re = '--repredict' in sys.argv
        res, msg = predict_goals_for_date(date, is_repredict=is_re)
        print(msg)
