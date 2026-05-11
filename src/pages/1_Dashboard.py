import streamlit as st
import json
import os
import pandas as pd
import sys

# 添加项目根目录到 Python 路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
from src.llm.predictor import LLMPredictor
from src.db.database import Database
from src.crawler.jingcai_crawler import JingcaiCrawler
from src.crawler.odds_crawler import OddsCrawler
from src.processor.data_fusion import DataFusion, build_leisu_crawler
import hashlib
from datetime import datetime, timedelta

import base64
from src.constants import AUTH_TOKEN_TTL
import time

def decode_auth_token(token):
    try:
        raw = base64.b64decode(token.encode('utf-8')).decode('utf-8')
        username, timestamp = raw.split('|')
        return username, int(timestamp)
    except:
        return None, 0

# ==========================================
# 路由守卫：尝试从 URL Params 恢复登录状态
# ==========================================
if "auth" in st.query_params and not st.session_state.get("logged_in", False):
    try:
        token = st.query_params["auth"]
        username, login_timestamp = decode_auth_token(token)
        
        if username and (int(time.time()) - login_timestamp <= AUTH_TOKEN_TTL):
            db = Database()
            user = db.get_user(username)
            db.close()
            
            if user and datetime.now() <= user.valid_until:
                st.session_state["logged_in"] = True
                st.session_state["username"] = user.username
                st.session_state["role"] = user.role
                st.session_state["valid_until"] = user.valid_until
    except Exception as e:
        pass

if not st.session_state.get("logged_in", False):
    st.warning("⚠️ 您尚未登录或会话已过期，请先登录！")
    if st.button("👉 返回登录页面"):
        st.switch_page("app.py")
    st.stop() # 停止渲染下方内容
# ==========================================

# 设置页面配置
st.set_page_config(
    page_title="泊松数据模型 - 看板",
    page_icon="⚽",
    layout="wide",
    initial_sidebar_state="expanded",
)

# 隐藏 Streamlit 默认的侧边栏页面导航
hide_pages_style = """
    <style>
        [data-testid="stSidebarNav"] {display: none;}
    </style>
"""
st.markdown(hide_pages_style, unsafe_allow_html=True)

# 加载数据
@st.cache_data(ttl=300) # 缓存5分钟
def load_data():
    data_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "data", "today_matches.json")
    if not os.path.exists(data_path):
        return []
    try:
        with open(data_path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        st.error(f"加载数据失败: {e}")
        return []

def save_data(matches):
    data_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "data", "today_matches.json")
    try:
        with open(data_path, "w", encoding="utf-8") as f:
            json.dump(matches, f, ensure_ascii=False, indent=2)
    except Exception as e:
        st.error(f"保存数据失败: {e}")


def _build_all_predictions_for_match(match, db_predictions=None):
    """合并数据库、缓存与最新 llm_prediction，确保页面展示拿到最新预测结果。"""
    all_predictions = dict(db_predictions or {})
    match_level_predictions = match.get("all_predictions") or {}
    all_predictions.update(match_level_predictions)

    latest_prediction = match.get("llm_prediction", "")
    if latest_prediction:
        try:
            period = LLMPredictor()._determine_prediction_period(match)
        except Exception:
            period = "final"
        all_predictions[period] = latest_prediction

    return all_predictions


def _resolve_primary_prediction_text(match, db_predictions=None):
    all_predictions = _build_all_predictions_for_match(match, db_predictions=db_predictions)
    if all_predictions:
        if "final" in all_predictions:
            return all_predictions["final"], all_predictions
        if "pre_12h" in all_predictions:
            return all_predictions["pre_12h"], all_predictions
        if "pre_24h" in all_predictions:
            return all_predictions["pre_24h"], all_predictions
        return list(all_predictions.values())[-1], all_predictions

    return match.get("llm_prediction", ""), all_predictions

def create_user_ui():
    st.sidebar.divider()
    st.sidebar.subheader("🛠️ Admin 管理面板")
    with st.sidebar.expander("➕ 新增/续期账号", expanded=False):
        with st.form("create_user_form"):
            new_username = st.text_input("用户名")
            new_password = st.text_input("密码", type="password")
            new_role = st.selectbox("角色", ["vip", "editor", "admin"], key="new_user_role")
            valid_days = st.number_input("有效期(天)", min_value=1, value=30)
            submit_btn = st.form_submit_button("保存账号")
            
            if submit_btn:
                if not new_username or not new_password:
                    st.error("用户名和密码不能为空")
                else:
                    try:
                        db = Database()
                        from src.db.database import User
                        user = db.session.query(User).filter_by(username=new_username).first()
                        valid_until = datetime.now() + timedelta(days=valid_days)
                        hashed_pw = hashlib.sha256(new_password.encode('utf-8')).hexdigest()
                        
                        if user:
                            user.password_hash = hashed_pw
                            user.role = new_role
                            user.valid_until = valid_until
                            st.success(f"更新成功！有效期至 {valid_until.strftime('%Y-%m-%d')}")
                        else:
                            new_user = User(
                                username=new_username,
                                password_hash=hashed_pw,
                                role=new_role,
                                valid_until=valid_until
                            )
                            db.session.add(new_user)
                            st.success(f"创建成功！有效期至 {valid_until.strftime('%Y-%m-%d')}")
                        db.session.commit()
                        db.close()
                    except Exception as e:
                        st.error(f"操作失败: {e}")

def main():
    user_info = f"👤 当前用户: {st.session_state['username']} ({st.session_state['role'].upper()})"
    if st.session_state.get('role') == 'vip' and 'valid_until' in st.session_state:
        # 兼容不同类型的有效时间格式
        valid_until = st.session_state['valid_until']
        if isinstance(valid_until, datetime):
            valid_date = valid_until.strftime('%Y-%m-%d')
        else:
            valid_date = str(valid_until).split(' ')[0]
        user_info += f"\n\n⏳ 到期时间: {valid_date}"
        
    st.sidebar.success(user_info)
    if st.sidebar.button("退出登录"):
        st.session_state["logged_in"] = False
        st.session_state["username"] = ""
        st.session_state["role"] = ""
        if "auth" in st.query_params:
            del st.query_params["auth"]
        st.switch_page("app.py")
        
    if st.session_state.get("role") == "admin":
        create_user_ui()
        
    st.title("⚽ 泊松数据模型与盘口分析看板")
    st.markdown("基于基本面与盘口异动分析，结合深度推理的竞彩推荐系统。")
    
    st.header("今日赛事概览")
    # TODO: 后续可在此处重构为 tabs 布局（包含：赛事概览、预测详情、智能串关、公众号推文、赛果复盘）
    
    matches = load_data()
    
    if not matches:
        st.warning("暂无今日赛事数据，请先运行数据抓取模块。")
        return
        
    # 侧边栏：赛事筛选
    st.sidebar.header("🔍 赛事筛选")
    leagues = sorted(list(set([m.get("league") for m in matches if m.get("league")])))
    selected_leagues = st.sidebar.multiselect("选择联赛", leagues, default=leagues)
    
    filtered_matches = [m for m in matches if m.get("league") in selected_leagues]
    
    st.sidebar.info(f"当前显示 {len(filtered_matches)} 场比赛")
    
    # 导航区
    st.sidebar.markdown("---")
    st.sidebar.header("🧭 功能导航")
    
    if st.sidebar.button("🏀 竞彩篮球预测", use_container_width=True):
        if "auth" in st.query_params:
            st.switch_page("pages/3_Basketball.py")
        else:
            try:
                raw_token = f"{st.session_state['username']}|{int(time.time())}"
                token = base64.b64encode(raw_token.encode('utf-8')).decode('utf-8')
                st.query_params["auth"] = token
            except:
                pass
            st.switch_page("pages/3_Basketball.py")
            
    if st.sidebar.button("🎯 足彩胜负彩 (14场)", use_container_width=True):
        if "auth" in st.query_params:
            st.switch_page("pages/4_ShengFuCai.py")
        else:
            try:
                raw_token = f"{st.session_state['username']}|{int(time.time())}"
                token = base64.b64encode(raw_token.encode('utf-8')).decode('utf-8')
                st.query_params["auth"] = token
            except:
                pass
            st.switch_page("pages/4_ShengFuCai.py")
        
    if st.sidebar.button("🔍 赛果复盘与模型优化", use_container_width=True):
        # 跳转时带上 auth token
        import urllib.parse
        if "auth" in st.query_params:
            st.switch_page("pages/2_Post_Mortem.py")
        else:
            # 如果是刚登录没刷新，构造一个token带过去
            try:
                raw_token = f"{st.session_state['username']}|{int(time.time())}"
                token = base64.b64encode(raw_token.encode('utf-8')).decode('utf-8')
                st.query_params["auth"] = token
            except:
                pass
            st.switch_page("pages/2_Post_Mortem.py")

    if st.session_state.get("role") == "admin":
        if st.sidebar.button("⚙️ 动态风控规则管理", use_container_width=True):
            if "auth" in st.query_params:
                st.switch_page("pages/5_Rule_Manager.py")
            else:
                try:
                    raw_token = f"{st.session_state['username']}|{int(time.time())}"
                    token = base64.b64encode(raw_token.encode('utf-8')).decode('utf-8')
                    st.query_params["auth"] = token
                except:
                    pass
                st.switch_page("pages/5_Rule_Manager.py")
    
    # 一键展开/收缩控制
    if 'expand_all' not in st.session_state:
        st.session_state.expand_all = False

    def toggle_expand():
        st.session_state.expand_all = not st.session_state.expand_all

    # 在左侧栏添加一键展开/收缩按钮
    st.sidebar.markdown("---")
    btn_text = "折叠所有赛事列表" if st.session_state.expand_all else "展开所有赛事列表"
    st.sidebar.button(btn_text, on_click=toggle_expand, use_container_width=True)
    
    # 在左侧边栏添加日志查看功能 (仅 Admin 可见)
    if st.session_state.get("role") == "admin":
        st.sidebar.markdown("---")
        if st.sidebar.button("📄 查看后台系统日志", use_container_width=True):
            st.session_state.show_logs = not st.session_state.get("show_logs", False)
            
        if st.session_state.get("show_logs", False):
            st.sidebar.markdown("### 🖥️ 系统日志")
            log_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "logs", "app.log")
            if os.path.exists(log_path):
                try:
                    with open(log_path, "r", encoding="utf-8") as f:
                        # 读取最后 50 行
                        lines = f.readlines()[-50:]
                        log_text = "".join(lines)
                        st.sidebar.text_area("最新 50 行日志", log_text, height=300, disabled=True)
                except Exception as e:
                    st.sidebar.error(f"读取日志失败: {e}")
            else:
                st.sidebar.info("暂无日志文件生成")
    
    # 主内容区：比赛列表
    # 全局重新预测功能 (仅 Admin)
    if st.session_state.get("role") == "admin":
        st.markdown("### ⚙️ 数据抓取与预测控制")
        
        # --- 新增的进球数专项控制台 ---
        st.markdown("#### 📊 进球数专项控制台")
        col_g1, col_g2, col_g3 = st.columns([1, 1, 1])
        with col_g1:
            target_date = st.date_input("选择目标日期 (用于进球数预测回写)", datetime.now().date())
            target_date_str = target_date.strftime('%Y-%m-%d')
            
        with col_g2:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("⚽ 按指定日期预测 (回写至'预测进球数')", use_container_width=True):
                with st.spinner(f"正在对 {target_date_str} 的比赛进行进球数推演并回写..."):
                    try:
                        import sys
                        import os
                        import importlib
                        sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
                        import scripts.batch_predict_goals as batch_predict_goals
                        importlib.reload(batch_predict_goals) # 强制重载模块
                        success, msg = batch_predict_goals.predict_goals_for_date(target_date_str, is_repredict=False)
                        if success:
                            st.success(msg)
                            st.toast("✅ " + msg)
                        else:
                            st.warning(msg)
                    except Exception as e:
                        st.error(f"操作失败: {e}")
                        
        with col_g3:
            st.markdown("<br>", unsafe_allow_html=True)
            if st.button("🔄 按指定日期重新预测 (回写至'重新预测')", use_container_width=True):
                with st.spinner(f"正在对 {target_date_str} 的比赛进行进球数重新推演并回写..."):
                    try:
                        import sys
                        import os
                        import importlib
                        sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
                        import scripts.batch_predict_goals as batch_predict_goals
                        importlib.reload(batch_predict_goals) # 强制重载模块以防 Streamlit 缓存旧代码
                        success, msg = batch_predict_goals.predict_goals_for_date(target_date_str, is_repredict=True)
                        if success:
                            st.success(msg)
                            st.toast("✅ " + msg)
                        else:
                            st.warning(msg)
                    except Exception as e:
                        st.error(f"操作失败: {e}")

        # 原有的统计报告按钮
        if st.button("� 更新进球数历史统计报告 (分析 Excel 实际进球分布)", use_container_width=True):
            with st.spinner("正在融合历史赛果与 Excel 数据，生成进球数概率分布报告..."):
                try:
                    import sys
                    import os
                    import importlib
                    sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
                    import scripts.temp_analyze as temp_analyze
                    importlib.reload(temp_analyze)
                    temp_analyze.analyze()
                    st.success("✅ 进球数历史统计报告已成功更新！可在 docs/goal_distribution_analysis.md 查看。大模型将使用最新数据进行推演。")
                except Exception as e:
                    st.error(f"统计失败: {e}")
        
        st.markdown("---")
        st.markdown("#### 🚀 全局抓取与主胜平负预测")
        
        # 目标日期选择
        col_d1, col_d2 = st.columns([1, 2])
        with col_d1:
            fetch_target_date = st.date_input("指定拉取日期", datetime.now().date(), key="fetch_date", 
                                              help="选择要拉取比赛的日期。500彩票网通常展示当前在售的3-4天比赛。")
        fetch_date_str = fetch_target_date.strftime('%Y-%m-%d')
        fetch_weekday_cn = ["周一","周二","周三","周四","周五","周六","周日"][fetch_target_date.weekday()]
        st.caption(f"目标日期: {fetch_date_str} ({fetch_weekday_cn})，将拉取该日编号匹配的比赛")
        
        time_options = ["全部时间段", "凌晨 (00:00-08:00)", "白天 (08:00-16:00)", "傍晚 (16:00-20:00)", "晚场 (20:00-24:00)", "自定义时间段"]
        
        col_t1, col_t2 = st.columns([1, 2])
        with col_t1:
            selected_time = st.selectbox("按开赛时间段过滤拉取", time_options)
            
        custom_start_h = 0
        custom_end_h = 24
        if selected_time == "自定义时间段":
            with col_t2:
                # 使用 slider 选择小时范围
                time_range = st.slider(
                    "选择开赛时间范围 (小时)",
                    min_value=0, max_value=24,
                    value=(0, 24),
                    step=1
                )
                custom_start_h, custom_end_h = time_range

        col_btn1, col_btn2, col_btn3 = st.columns([1, 1, 1])
        with col_btn1:
            if st.button("🚀 重新拉取数据并全局预测 (仅全场)", type="primary", use_container_width=True):
                time_msg = selected_time if selected_time != "自定义时间段" else f"{custom_start_h}:00 - {custom_end_h}:00"
                with st.spinner(f"正在从500网拉取 {fetch_date_str} {time_msg} 的赛事数据，并调用大模型重新预测..."):
                    # 1. 重新抓取数据（传入目标日期）
                    st.toast(f"正在抓取 {fetch_date_str} ({fetch_weekday_cn}) 竞彩赛事列表...")
                    jingcai_crawler = JingcaiCrawler()
                    new_matches = jingcai_crawler.fetch_today_matches(target_date=fetch_target_date)
                    
                    if selected_time != "全部时间段":
                        if selected_time == "自定义时间段":
                            start_h, end_h = custom_start_h, custom_end_h
                        else:
                            ranges = {
                                "凌晨 (00:00-08:00)": (0, 8),
                                "白天 (08:00-16:00)": (8, 16),
                                "傍晚 (16:00-20:00)": (16, 20),
                                "晚场 (20:00-24:00)": (20, 24)
                            }
                            start_h, end_h = ranges[selected_time]
                            
                        filtered_new = []
                        for m in new_matches:
                            try:
                                hour = int(m.get('match_time', '').split(' ')[1].split(':')[0])
                                if start_h <= hour < end_h:
                                    filtered_new.append(m)
                            except:
                                filtered_new.append(m) # 解析失败的也保留，防止漏掉
                        new_matches = filtered_new

                    if not new_matches:
                        st.error(f"未能抓取到 {fetch_date_str} {time_msg} 的比赛数据，请确认500彩票网该日期有在售比赛。")
                    else:
                        st.toast("正在抓取盘口与基本面数据并融合...")
                        odds_crawler = OddsCrawler()
                        data_fusion = DataFusion()
                        leisu = build_leisu_crawler(headless=True)
                        merged_matches = data_fusion.merge_data(new_matches, odds_crawler, leisu_crawler=leisu)
                        if leisu:
                            try:
                                leisu.close()
                            except Exception:
                                pass
                        
                        # 补充从 Excel 读取进球数相关数据
                        try:
                            import pandas as pandas_lib
                            excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "docs", "foot_prediction.xlsx")
                            if os.path.exists(excel_path):
                                xl = pandas_lib.ExcelFile(excel_path)
                                
                                for m in merged_matches:
                                    match_num = m.get('match_num')
                                    for sheet_name in reversed(xl.sheet_names):
                                        df_excel = pandas_lib.read_excel(xl, sheet_name=sheet_name)
                                        df_excel.columns = df_excel.columns.str.strip()
                                        if '编码' not in df_excel.columns:
                                            continue
                                        df_excel['编码'] = df_excel['编码'].astype(str).str.strip()
                                        row = df_excel[df_excel['编码'] == match_num]
                                        if not row.empty:
                                            pan = row.iloc[0].get('进球盘口', '')
                                            diff = row.iloc[0].get('预测差异百分比', '')
                                            trend = row.iloc[0].get('倾向', '')
                                            m['goals_pan'] = str(pan).strip() if pandas_lib.notna(pan) else ""
                                            m['goals_diff_percent'] = str(diff).strip() if pandas_lib.notna(diff) else ""
                                            m['goals_trend'] = str(trend).strip() if pandas_lib.notna(trend) else ""
                                            break
                        except Exception as e:
                            st.warning(f"读取 Excel 数据补充失败: {e}")
                        
                        st.toast("正在全局重新调用大模型预测(仅全场)...")
                        import sys
                        import os
                        from src.llm.predictor import LLMPredictor
                        predictor = LLMPredictor()
                        
                        db = Database()
                        total_count = len(merged_matches)
                        for m in merged_matches:
                            other_matches_context = [om for om in merged_matches if om.get('match_num') != m.get('match_num')]
                            res, period = predictor.predict(m, total_matches_count=total_count, other_matches_context=other_matches_context)
                            m["llm_prediction"] = res
                            if "all_predictions" not in m:
                                m["all_predictions"] = {}
                            m["all_predictions"][period] = res
                            
                            db.save_prediction(m, period)
                        db.close()
                        
                        # 覆盖保存
                        load_data.clear()
                        save_data(merged_matches)
                        st.success("最新赛事数据已拉取并全局全场预测完成！")
                        time.sleep(1)
                        st.rerun()
                        
        with col_btn2:
            if st.button("🚀 仅对当前数据全局重新预测 (仅全场)", use_container_width=True):
                with st.spinner("正在全局重新调用大模型进行全场预测，请耐心等待..."):
                    from src.llm.predictor import LLMPredictor
                    from src.processor.data_fusion import inject_leisu_data
                    predictor = LLMPredictor()
                    leisu = build_leisu_crawler(headless=True)

                    db = Database()
                    total_count = len(matches)
                    for m in matches:
                        if leisu:
                            inject_leisu_data(m, leisu)
                        other_matches_context = [om for om in matches if om.get('match_num') != m.get('match_num')]
                        res, period = predictor.predict(m, total_matches_count=total_count, other_matches_context=other_matches_context)
                        m["llm_prediction"] = res
                        if "all_predictions" not in m:
                            m["all_predictions"] = {}
                        m["all_predictions"][period] = res
                        db.save_prediction(m, period)

                    if leisu:
                        try: leisu.close()
                        except Exception: pass
                    db.close()
                    load_data.clear()
                    save_data(matches)
                    st.success("全局全场预测完成！")
                    time.sleep(1)
                    st.rerun()

        with col_btn3:
            if st.button("⚽ 对当前数据全局执行进球数推演并回写", use_container_width=True):
                with st.spinner("正在全局调用大模型进行进球数专项预测并回写Excel..."):
                    import sys
                    import os
                    from src.llm.goals_predictor import GoalsPredictor
                    goals_predictor = GoalsPredictor()
                    
                    # 1. 先从 Excel 中实时读取最新的进球盘口等数据，注入到 matches 中
                    try:
                        import pandas as pandas_lib
                        excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "docs", "foot_prediction.xlsx")
                        if os.path.exists(excel_path):
                            xl = pandas_lib.ExcelFile(excel_path)
                            
                            for m in matches:
                                match_num = m.get('match_num')
                                # 遍历所有 sheet 找这行
                                for sheet_name in reversed(xl.sheet_names):
                                    df_excel = pandas_lib.read_excel(xl, sheet_name=sheet_name)
                                    df_excel.columns = df_excel.columns.str.strip()
                                    if '编码' not in df_excel.columns:
                                        continue
                                    df_excel['编码'] = df_excel['编码'].astype(str).str.strip()
                                    row = df_excel[df_excel['编码'] == match_num]
                                    if not row.empty:
                                        pan = row.iloc[0].get('进球盘口', '')
                                        diff = row.iloc[0].get('预测差异百分比', '')
                                        trend = row.iloc[0].get('倾向', '')
                                        
                                        m['goals_pan'] = str(pan).strip() if pandas_lib.notna(pan) else ""
                                        m['goals_diff_percent'] = str(diff).strip() if pandas_lib.notna(diff) else ""
                                        m['goals_trend'] = str(trend).strip() if pandas_lib.notna(trend) else ""
                                        # 记录在哪个 sheet 找到的，方便后续回写
                                        m['excel_sheet_name'] = sheet_name
                                        break
                    except Exception as e:
                        st.warning(f"读取 Excel 数据失败: {e}")

                    excel_updates = {}
                    processed_count = 0
                    
                    for m in matches:
                        # 只要有该比赛的Excel记录就可以推演，我们稍微放宽一下判断条件
                        if m.get('goals_pan') or m.get('goals_diff_percent') or m.get('goals_trend'):
                            processed_count += 1
                            goals_pred, _ = goals_predictor.predict(m)
                            
                            try:
                                res_dict = json.loads(goals_pred)
                                stat_goals = res_dict.get('statistical_goals')
                                fund_report = res_dict.get('fundamental_report', '')
                            except Exception as e:
                                stat_goals = None
                                fund_report = goals_pred

                            m["goals_prediction_special"] = fund_report
                            
                            # 提取统计数据洞察打印到控制台
                            insight_text = ""
                            import re
                            insight_match = re.search(r'【统计数据洞察】\s*(.*?)(?=\n\n|\n\*\*|$)', fund_report, re.DOTALL)
                            if insight_match:
                                insight_text = insight_match.group(1).strip().replace('\n', ' | ')
                            print(f"全局推演 {m.get('match_num')} 统计预测: {stat_goals}")
                            if insight_text:
                                print(f"  └─ 洞察: {insight_text}")
                            
                            if "预测失败" in fund_report:
                                st.error(f"比赛 {m.get('match_num')} 进球数预测失败: {fund_report}")
                                
                            # 提取进球数
                            import re
                            clean_pred = re.sub(r'\*+', '', fund_report).strip()
                            match_goals = re.search(r'【进球数预测】\s*([^\n]+)', clean_pred)
                            extracted_goals = None
                            if match_goals:
                                extracted_goals = match_goals.group(1).strip()
                            else:
                                match_goals = re.search(r'进球数预测.*?\n([^\n]+)', clean_pred)
                                if match_goals:
                                    extracted_goals = match_goals.group(1).strip()
                                        
                            if stat_goals or extracted_goals:
                                excel_updates[m.get('match_num')] = {
                                    'stat_goals': stat_goals,
                                    'fund_goals': extracted_goals
                                }
                                
                    if processed_count == 0:
                        st.warning("⚠️ 未能从 Excel 中匹配到任何今天比赛的进球数盘口等数据，请检查 Excel 的【编码】列是否与当前赛事列表的编号（如 周一001）一致，且填写了进球盘口等数据。")
                        return
                        
                    # 批量回写 Excel
                    if excel_updates:
                        try:
                            import os
                            from openpyxl import load_workbook
                            excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "docs", "foot_prediction.xlsx")
                            if os.path.exists(excel_path):
                                wb = load_workbook(excel_path)
                                
                                # 按 sheet 分组回写
                                updates_by_sheet = {}
                                for m in matches:
                                    match_num = m.get('match_num')
                                    if match_num in excel_updates and 'excel_sheet_name' in m:
                                        sheet_name = m['excel_sheet_name']
                                        if sheet_name not in updates_by_sheet:
                                            updates_by_sheet[sheet_name] = {}
                                        updates_by_sheet[sheet_name][match_num] = excel_updates[match_num]
                                        
                                for sheet_name, sheet_updates in updates_by_sheet.items():
                                    if sheet_name in wb.sheetnames:
                                        ws = wb[sheet_name]
                                        code_col_idx, pred_goals_col_idx, fund_goals_col_idx = None, None, None
                                        
                                        for col_idx, cell in enumerate(ws[1], 1):
                                            if cell.value and str(cell.value).strip() == '编码':
                                                code_col_idx = col_idx
                                            elif cell.value and '预测进球数' in str(cell.value).strip() and '基本面' not in str(cell.value).strip():
                                                pred_goals_col_idx = col_idx
                                            elif cell.value and '基本面预测进球数' in str(cell.value).strip():
                                                fund_goals_col_idx = col_idx
                                                
                                        if not fund_goals_col_idx:
                                            fund_goals_col_idx = ws.max_column + 1
                                            ws.cell(row=1, column=fund_goals_col_idx).value = '基本面预测进球数'
                                        
                                        if code_col_idx and pred_goals_col_idx:
                                            for row_idx in range(2, ws.max_row + 1):
                                                match_num = str(ws.cell(row=row_idx, column=code_col_idx).value).strip()
                                                if match_num in sheet_updates:
                                                    upd = sheet_updates[match_num]
                                                    
                                                    if upd['stat_goals']:
                                                        ws.cell(row=row_idx, column=pred_goals_col_idx).value = upd['stat_goals']
                                                    elif upd['fund_goals']:
                                                        ws.cell(row=row_idx, column=pred_goals_col_idx).value = upd['fund_goals']
                                                        
                                                    if upd['fund_goals']:
                                                        ws.cell(row=row_idx, column=fund_goals_col_idx).value = upd['fund_goals']
                                                    
                                wb.save(excel_path)
                                st.toast("✅ Excel 全局回写成功！")
                        except Exception as e:
                            st.error(f"批量回写 Excel 失败: {e}")
                    
                    load_data.clear()
                    save_data(matches)
                    st.success("全局进球数预测完成！")
                    time.sleep(1)
                    st.rerun()

        # ========== 历史数据拉取（单独按钮，仅拉取不预测，用于复盘） ==========
        st.markdown("---")
        st.markdown("#### 📦 历史数据拉取（仅拉取已完赛数据，不预测）")
        
        col_h1, col_h2 = st.columns([1, 2])
        with col_h1:
            history_date = st.date_input("选择历史日期", datetime.now().date() - timedelta(days=1), key="history_date",
                                         help="从500彩票网历史页面拉取指定日期的已完赛比赛数据（含赔率和比分）")
        history_date_str = history_date.strftime('%Y-%m-%d')
        history_weekday = ["周一","周二","周三","周四","周五","周六","周日"][history_date.weekday()]
        
        if st.button("📥 拉取历史比赛数据（仅入库，不预测）", use_container_width=True):
            with st.spinner(f"正在从500网历史页面拉取 {history_date_str} ({history_weekday}) 的比赛数据..."):
                jingcai_crawler = JingcaiCrawler()
                history_matches = jingcai_crawler.fetch_history_matches(history_date)
                
                if not history_matches:
                    st.error(f"未能拉取到 {history_date_str} 的历史比赛数据。")
                else:
                    st.toast(f"已拉取 {len(history_matches)} 场比赛，正在融合盘口数据...")
                    
                    # 融合亚盘数据
                    try:
                        odds_crawler = OddsCrawler()
                        data_fusion = DataFusion()
                        history_matches = data_fusion.merge_data(history_matches, odds_crawler)
                    except Exception as e:
                        st.warning(f"盘口数据融合失败（不影响基本信息）: {e}")
                    
                    # 保存到数据库（使用原始SQL避免依赖预测字段）
                    db = Database()
                    saved = 0
                    for m in history_matches:
                        try:
                            raw_json = json.dumps(m, ensure_ascii=False)
                            match_time = m.get('match_time', '')
                            from sqlalchemy import text
                            db.session.execute(text("""
                                INSERT OR REPLACE INTO match_predictions 
                                (fixture_id, match_num, league, home_team, away_team, match_time, 
                                 prediction_text, prediction_period, raw_data, actual_score, created_at)
                                VALUES (:fid, :mn, :lg, :ht, :at, :mt, '', 'historical', :raw, :sc, datetime('now'))
                            """), {
                                "fid": m.get('fixture_id'), "mn": m.get('match_num'),
                                "lg": m.get('league', ''), "ht": m.get('home_team', ''),
                                "at": m.get('away_team', ''), "mt": match_time,
                                "raw": raw_json, "sc": m.get('actual_score', '')
                            })
                            saved += 1
                        except Exception as e:
                            st.warning(f"保存 {m.get('match_num')} 失败: {e}")
                    db.session.commit()
                    db.close()
                    
                    st.success(f"✅ 成功拉取并入库 {saved}/{len(history_matches)} 场历史比赛数据！可前往「赛果复盘与模型优化」页面进行复盘。")
                    st.info(f"日期: {history_date_str} ({history_weekday})，包含 {saved} 场比赛的赔率和比分。")

    # 汇总预测结果看板
    st.markdown("---")
    st.subheader("📋 今日赛事预测汇总")
    
    # 收集所有预测结果
    summary_data = []
    htft_summary_data = []
    goals_summary_data = []
    db = Database()
    db_preds_cache = {}
    for match in filtered_matches:
        # 1. 收集全场赛果预测
        fixture_id = match.get("fixture_id")
        if fixture_id:
            if fixture_id not in db_preds_cache:
                all_predictions = {}
                preds = db.get_all_predictions_by_fixture(fixture_id)
                for pred in preds:
                    all_predictions[pred.prediction_period] = pred.prediction_text
                db_preds_cache[fixture_id] = all_predictions
            db_all_preds = db_preds_cache.get(fixture_id) or {}
        else:
            db_all_preds = {}

        prediction_text, merged_all_preds = _resolve_primary_prediction_text(match, db_predictions=db_all_preds)
        if merged_all_preds:
            match["all_predictions"] = merged_all_preds
            
        # 尝试提取预测结果以显示在标题栏
        recommendation = "待分析"
        pred_color = "gray"
        
        if prediction_text:
            from src.llm.predictor import LLMPredictor
            details = LLMPredictor.parse_prediction_details(prediction_text)
            
            # 用于标题栏的推荐提取
            rec_nspf = details.get('recommendation_nspf', '')
            if rec_nspf:
                recommendation = rec_nspf
                pred_color = "green" if "胜" in recommendation else ("blue" if "平" in recommendation else "red")
            else:
                recommendation = "已分析 (需点开查看)"
                pred_color = "black"
            
            # 提取置信度分数进行排序
            confidence_str = details.get('confidence', '0')
            import re
            conf_match = re.search(r'\d+', confidence_str)
            conf_score = int(conf_match.group()) if conf_match else 0
            
            # 提取竞彩赔率供串关方案参考
            match_odds = match.get('odds', {})
            nspf_sp = match_odds.get('nspf', [])  # 不让球赔率 [胜, 平, 负]
            spf_sp = match_odds.get('spf', [])    # 让球赔率 [胜, 平, 负]
            rangqiu = match_odds.get('rangqiu', '0')
            
            summary_data.append({
                "编号": match.get('match_num', ''),
                "赛事": match.get('league', ''),
                "主队": match.get('home_team', ''),
                "客队": match.get('away_team', ''),
                "开赛时间": match.get('match_time', ''),
                "竞彩推荐": details.get('recommendation', '无'),
                "竞彩推荐(不让球)": details.get('recommendation_nspf', '无'),
                "竞彩让球推荐": details.get('recommendation_rq', '无'),
                "不让球赔率(胜/平/负)": nspf_sp if len(nspf_sp) == 3 else [],
                "让球赔率(胜/平/负)": spf_sp if len(spf_sp) == 3 else [],
                "让球数": rangqiu,
                "胜平负置信度": confidence_str,
                "比分参考": details.get('score', '无'),
                "基础理由": details.get('reason', '无'),
                "_sort_score": conf_score
            })
            
        # 1.5 收集进球数专项预测
        goals_pred_text = match.get("goals_prediction_special", "")
        if goals_pred_text:
            import re
            extracted_goals = "无"
            # 支持匹配换行和不同符号
            clean_pred = re.sub(r'\*+', '', goals_pred_text).strip()
            match_goals = re.search(r'【进球数预测】\s*([^\n]+)', clean_pred)
            if match_goals:
                extracted_goals = match_goals.group(1).strip()
            else:
                # 尝试更宽泛的匹配，可能没有加粗或者冒号
                match_goals = re.search(r'进球数预测.*?\n([^\n]+)', clean_pred)
                if match_goals:
                    extracted_goals = match_goals.group(1).strip()
                    
            # 计算联赛特征组
            league_name = match.get('league', '')
            cluster = 'D组：其他未分类联赛'
            group_a = ['澳超', '挪超', '荷甲', '瑞超', '日职', '美职足', '沙特职业联赛', '芬兰超级联赛']
            group_b = ['意甲', '西乙', '法乙', '阿甲', '葡超', '韩职', '英冠']
            group_c = ['英超', '德甲', '西甲', '法甲', '德乙']
            
            for g in group_a:
                if g in league_name: cluster = 'A组：开放大开大合型'
            for g in group_b:
                if g in league_name: cluster = 'B组：严密防守型'
            for g in group_c:
                if g in league_name: cluster = 'C组：主流均衡型'
                
            goals_summary_data.append({
                "编号": match.get('match_num', ''),
                "赛事": league_name,
                "参考特征组": cluster,
                "主队": match.get('home_team', ''),
                "客队": match.get('away_team', ''),
                "开赛时间": match.get('match_time', ''),
                "进球数盘口(机构)": match.get('goals_pan', '无'),
                "预测差异百分比": match.get('goals_diff_percent', '无'),
                "倾向": match.get('goals_trend', '无'),
                "AI预测进球数": extracted_goals,
                "_sort_score": conf_score if 'conf_score' in locals() else 0
            })
            
        # 2. 收集半全场预测
        htft_text = match.get("htft_prediction", "")
        if htft_text:
            # 简单解析半全场推荐结果 (从 Markdown 中提取)
            htft_rec = "无"
            htft_conf = "0"
            import re
            
            # 支持更多可能的冒号格式和不带中括号的格式，以及跨行情况
            rec_match = re.search(r'半全场单关推荐[^\n]*?[:：]\s*\[?(.*?)\]?(?=\n|$)', htft_text)
            if rec_match:
                htft_rec = rec_match.group(1).replace('[', '').replace(']', '').replace('**', '').strip()
                
            # 匹配各种可能的数字组合，包括：**35**，【55/100】，68/100，**70**，等
            conf_match = re.search(r'半全场置信度[^\d]*(\d+)', htft_text)
            if conf_match:
                htft_conf_val = int(conf_match.group(1))
                # 如果是建议放弃，大模型给出的置信度通常是“放弃该玩法的置信度”
                # 这种情况下，我们对于打出“平胜/平负”的真实置信度应该反过来（比如放弃置信度80，真实打出置信度就是20）
                # 考虑到周五008的推荐被解析为“无”或者其他情况，我们直接在全文中搜索“放弃”关键词
                if '放弃' in htft_rec or '放弃' in htft_text[htft_text.find('最终预测'):]:
                    htft_conf_val = max(0, 100 - htft_conf_val)
                htft_conf = str(htft_conf_val)
                
            htft_summary_data.append({
                "编号": match.get('match_num', ''),
                "赛事": match.get('league', ''),
                "主队": match.get('home_team', ''),
                "客队": match.get('away_team', ''),
                "开赛时间": match.get('match_time', ''),
                "半全场推荐": htft_rec,
                "置信度": htft_conf,
                "_sort_score": int(htft_conf) if htft_conf.isdigit() else 0
            })
            
    db.close()

    if summary_data or htft_summary_data or goals_summary_data:
        with st.expander("📊 展开查看今日赛事预测汇总列表", expanded=True):
            # 使用 Tabs 分开展示 全场预测 和 半全场预测 的汇总列表
            sum_tab1, sum_tab2, sum_tab3 = st.tabs(["🎯 全场赛果推荐汇总", "🌗 半全场(平胜/平负)推荐汇总", "⚽ 进球数专项预测汇总"])
            
            with sum_tab1:
                if summary_data:
                    df_summary = pd.DataFrame(summary_data)
                    df_summary = df_summary.sort_values(by='_sort_score', ascending=False).drop(columns=['_sort_score'])
                    cols_order = ["编号", "赛事", "主队", "客队", "开赛时间", "竞彩推荐", "胜平负置信度", "比分参考", "基础理由"]
                    df_summary = df_summary[cols_order]
                    st.dataframe(df_summary, use_container_width=True, hide_index=True)
                else:
                    st.info("暂无全场预测数据")
                    
            with sum_tab2:
                if htft_summary_data:
                    df_htft = pd.DataFrame(htft_summary_data)
                    # 过滤掉放弃玩法的比赛，只显示有推荐价值的
                    df_htft_filtered = df_htft[~df_htft['半全场推荐'].str.contains('放弃', na=False, case=False)]
                    
                    if not df_htft_filtered.empty:
                        df_htft_filtered = df_htft_filtered.sort_values(by='_sort_score', ascending=False).drop(columns=['_sort_score'])
                        st.dataframe(df_htft_filtered, use_container_width=True, hide_index=True)
                    else:
                        st.info("今日暂无符合【平胜/平负】单关模式的优质推荐赛事。")
                    
                    # 可选：提供一个复选框查看所有（包括放弃的）
                    if st.checkbox("显示所有赛事（包含建议放弃半全场玩法的比赛）"):
                        df_htft_all = df_htft.sort_values(by='_sort_score', ascending=False).drop(columns=['_sort_score'])
                        st.dataframe(df_htft_all, use_container_width=True, hide_index=True)
                else:
                    st.info("暂无半全场专项预测数据")
                    
            with sum_tab3:
                if goals_summary_data:
                    df_goals = pd.DataFrame(goals_summary_data)
                    df_goals = df_goals.sort_values(by='_sort_score', ascending=False).drop(columns=['_sort_score'])
                    cols_order_g = ["编号", "赛事", "参考特征组", "主队", "客队", "开赛时间", "进球数盘口(机构)", "预测差异百分比", "倾向", "AI预测进球数"]
                    df_goals = df_goals[cols_order_g]
                    st.dataframe(df_goals, use_container_width=True, hide_index=True)
                else:
                    st.info("今日暂无进球数专项预测数据。请先执行全局进球数推演。")
        
        # ---------------- 智能串关生成功能 ----------------
        st.markdown("### 🎫 智能串关方案推荐")
        
        # 为了串关能用到最新的进球数专项预测结果，我们在传递数据前融合一下
        combined_summary_data = []
        for s in summary_data:
            # 找到对应的进球数专项数据
            goal_data = next((g for g in goals_summary_data if g["编号"] == s["编号"]), None)
            if goal_data and goal_data["AI预测进球数"] != "无":
                s_copy = s.copy()
                s_copy["AI预测进球数"] = goal_data["AI预测进球数"]
                combined_summary_data.append(s_copy)
            else:
                combined_summary_data.append(s)

        def _parse_kickoff_time(value):
            if not value:
                return None
            text = str(value).strip().replace("T", " ")
            if "." in text:
                text = text.split(".", 1)[0]
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M"):
                try:
                    return datetime.strptime(text, fmt)
                except ValueError:
                    continue
            try:
                return datetime.fromisoformat(text)
            except ValueError:
                return None

        if "parlay_time_filter_enabled" not in st.session_state:
            st.session_state.parlay_time_filter_enabled = False
        if "parlay_time_start_date" not in st.session_state:
            st.session_state.parlay_time_start_date = datetime.now().date()
        if "parlay_time_end_date" not in st.session_state:
            st.session_state.parlay_time_end_date = datetime.now().date()
        if "parlay_time_start_time" not in st.session_state:
            st.session_state.parlay_time_start_time = datetime.strptime("18:00", "%H:%M").time()
        if "parlay_time_end_time" not in st.session_state:
            st.session_state.parlay_time_end_time = datetime.strptime("22:00", "%H:%M").time()

        enable_filter = st.checkbox("限定比赛时间范围", key="parlay_time_filter_enabled")
        filter_error = None
        if enable_filter:
            col_t1, col_t2 = st.columns(2)
            with col_t1:
                st.date_input("开始日期", key="parlay_time_start_date")
                st.time_input("开始时间", key="parlay_time_start_time")
            with col_t2:
                st.date_input("结束日期", key="parlay_time_end_date")
                st.time_input("结束时间", key="parlay_time_end_time")

            start_dt = datetime.combine(st.session_state.parlay_time_start_date, st.session_state.parlay_time_start_time)
            end_dt = datetime.combine(st.session_state.parlay_time_end_date, st.session_state.parlay_time_end_time)
            if end_dt <= start_dt:
                filter_error = "时间范围无效：结束时间必须大于开始时间。"
                st.error(filter_error)

            parlay_summary_data = []
            if not filter_error:
                for row in combined_summary_data:
                    kickoff = _parse_kickoff_time(row.get("开赛时间", ""))
                    if kickoff is None:
                        continue
                    if start_dt <= kickoff <= end_dt:
                        parlay_summary_data.append(row)
            st.caption(f"串关候选场次：{len(parlay_summary_data)} / {len(combined_summary_data)}")
        else:
            parlay_summary_data = combined_summary_data

        # 用于保存生成的串子单结果
        if "generated_parlays" not in st.session_state:
            st.session_state.generated_parlays = ""
        if "previous_parlays" not in st.session_state:
            st.session_state.previous_parlays = ""
        if "parlays_comparison" not in st.session_state:
            st.session_state.parlays_comparison = ""
        if "generated_parlays_structured" not in st.session_state:
            st.session_state.generated_parlays_structured = []
        if "previous_parlays_structured" not in st.session_state:
            st.session_state.previous_parlays_structured = []

        def render_structured_parlays(plans):
            if not plans:
                return False

            for plan in plans:
                payout = plan.get("payout", {})
                with st.container(border=True):
                    title_col, target_col = st.columns([2, 1])
                    with title_col:
                        st.markdown(f"### 方案{plan.get('plan_code', '')}：{plan.get('plan_name', '')}")
                    with target_col:
                        st.info(plan.get("target_status", ""))

                    metric_col1, metric_col2, metric_col3 = st.columns(3)
                    metric_col1.metric("净回报区间", f"{payout.get('net_min', 0):.2f} ~ {payout.get('net_max', 0):.2f}")
                    metric_col2.metric("总注数", str(payout.get("notes_count", 0)))
                    metric_col3.metric("理论赔率", f"{payout.get('min_product', 0):.2f} ~ {payout.get('max_product', 0):.2f}")

                    st.caption(plan.get("role_desc", ""))
                    st.write(plan.get("logic_summary", ""))

                    match_tabs = st.tabs(["入选场次", "赔率推演", "备选替换场"])

                    with match_tabs[0]:
                        for match in plan.get("matches", []):
                            st.markdown(
                                f"**[{match.get('match_id', '')}] {match.get('home_team', '')} VS {match.get('away_team', '')}**"
                            )
                            st.write(f"推荐：{match.get('selection_text', '')}")
                            st.write(
                                f"置信度：{match.get('confidence', 0)} | 标签：{' / '.join(match.get('tags', [])) or '无'} | 进球数参考：{match.get('goals_ref', '无')}"
                            )
                            st.write(f"入选理由：{'；'.join(match.get('selection_reasons', [])) or '无'}")
                            st.write(f"风险提示：{'；'.join(match.get('risk_notes', [])) or '无'}")
                            st.markdown("---")

                    with match_tabs[1]:
                        for odds_line in payout.get("odds_lines", []):
                            st.write(
                                f"{odds_line.get('match_id', '')} ({odds_line.get('selection', '')})：{odds_line.get('odds_text', '')}"
                            )
                        if payout.get("odds_lines"):
                            st.write(
                                f"注数计算：{' × '.join(payout.get('notes_factors', []))} = {payout.get('notes_count', 0)} 注"
                            )
                            st.write(
                                f"理论最低赔率：{' × '.join(payout.get('min_factors', []))} = {payout.get('min_product', 0):.2f}"
                            )
                            st.write(
                                f"理论最高赔率：{' × '.join(payout.get('max_factors', []))} = {payout.get('max_product', 0):.2f}"
                            )
                            st.write(
                                f"真实净回报：最低 {payout.get('net_min', 0):.2f} 倍 ~ 最高 {payout.get('net_max', 0):.2f} 倍"
                            )
                        else:
                            st.write("当前未凑齐满足约束的两场组合，请结合备选替换场人工调整。")

                    with match_tabs[2]:
                        alternative = plan.get("alternative")
                        if alternative:
                            st.write(
                                f"[{alternative.get('match_id', '')}] {alternative.get('home_team', '')} VS {alternative.get('away_team', '')}"
                            )
                            st.write(f"推荐：{alternative.get('selection_text', '')}")
                            st.write(f"替换理由：{'；'.join(alternative.get('selection_reasons', [])) or '无'}")
                        else:
                            st.write("暂无可用备选替换场。")
            return True
            
        # 尝试从数据库加载今日串关历史（只在页面首次加载且 session state 为空时）
        today_date = datetime.now().strftime("%Y-%m-%d")
        if not st.session_state.generated_parlays:
            db = Database()
            saved_parlay_record = db.get_parlays_by_date(today_date)
            if saved_parlay_record:
                st.session_state.generated_parlays = saved_parlay_record.current_parlay or ""
                st.session_state.previous_parlays = saved_parlay_record.previous_parlay or ""
                st.session_state.parlays_comparison = saved_parlay_record.comparison_text or ""
            db.close()
            
        col_btn_parlay, _ = st.columns([1, 2])
        with col_btn_parlay:
            btn_text = "🔄 重新评估今日三套方案" if st.session_state.generated_parlays else "🎯 AI 智能生成三套实战串子单"
            if st.button(btn_text, type="primary", use_container_width=True):
                with st.spinner("操盘手 AI 正在深度分析所有赛事并为您组合串子单，请稍候..."):
                    import sys
                    import os
                    from src.llm.predictor import LLMPredictor
                    predictor = LLMPredictor()
                    if enable_filter and filter_error:
                        st.error("时间范围无效，无法生成串关方案。")
                        st.stop()
                    if len(parlay_summary_data) < 2:
                        st.error("候选场次不足（至少需要 2 场）")
                        st.stop()
                    payload = predictor.generate_parlays_payload(parlay_summary_data)
                    new_parlays = payload["markdown"]
                    
                    # 如果已经有老的方案，则保存为历史并进行对比
                    if st.session_state.generated_parlays:
                        st.session_state.previous_parlays = st.session_state.generated_parlays
                        st.session_state.previous_parlays_structured = st.session_state.generated_parlays_structured
                        
                        # 只有当新旧方案不完全相同时才去对比
                        if new_parlays != st.session_state.previous_parlays:
                            with st.spinner("发现新方案与旧方案不一致！风控总监 AI 正在为您深度对比分析..."):
                                comparison = predictor.compare_parlays(
                                    st.session_state.previous_parlays, 
                                    new_parlays
                                )
                                st.session_state.parlays_comparison = comparison
                        else:
                            st.session_state.parlays_comparison = "两次生成的方案完全一致，无需对比。"
                            
                    st.session_state.generated_parlays = new_parlays
                    st.session_state.generated_parlays_structured = payload["plans"]
                    
                    # 持久化保存到数据库
                    db = Database()
                    db.save_parlays(
                        target_date=today_date,
                        current_parlay=st.session_state.generated_parlays,
                        previous_parlay=st.session_state.previous_parlays,
                        comparison_text=st.session_state.parlays_comparison
                    )
                    db.close()
                    
        # 展示串关结果
        if st.session_state.generated_parlays:
            with st.expander("🎫 AI 智能串关方案与深度对比", expanded=True):
                if st.session_state.previous_parlays and st.session_state.parlays_comparison:
                    st.success("✅ 发现多套方案，已为您生成深度对比！")
                    
                    # 使用 Tabs 来展示最新、历史和对比
                    tab_new, tab_comp, tab_old = st.tabs(["🌟 最新生成方案", "⚖️ AI 深度对比分析", "🕰️ 前次历史方案"])
                    
                    with tab_new:
                        if not render_structured_parlays(st.session_state.generated_parlays_structured):
                            st.markdown(st.session_state.generated_parlays)
                        
                    with tab_comp:
                        st.info("💡 **风控提示**：AI 针对两次不同的选场给出了深度分析，建议结合阅读。")
                        st.markdown(st.session_state.parlays_comparison)
                        
                    with tab_old:
                        if not render_structured_parlays(st.session_state.previous_parlays_structured):
                            st.markdown(st.session_state.previous_parlays)
                else:
                    st.success("✅ 串子单生成完毕！请参考：")
                    if not render_structured_parlays(st.session_state.generated_parlays_structured):
                        st.markdown(st.session_state.generated_parlays)
            
    else:
        st.info("暂无预测数据，请先运行预测。")
    
    st.markdown("---")
    st.subheader("📝 赛事详细分析列表")

    # ========== 顶层：处理单场重新预测请求（防循环内重复触发） ==========
    _pending_actions = []
    for _m in filtered_matches:
        _mn = _m.get('match_num')
        if st.session_state.pop(f"repredict_pending_{_mn}", None):
            _pending_actions.append(('repredict', _m))
        if st.session_state.pop(f"repredict_htft_pending_{_mn}", None):
            _pending_actions.append(('htft', _m))
        if st.session_state.pop(f"repredict_goals_pending_{_mn}", None):
            _pending_actions.append(('goals', _m))

    if _pending_actions:
        _action_type, _target_match = _pending_actions[0]

        # 注入雷速体育数据
        leisu = None
        try:
            from src.processor.data_fusion import inject_leisu_data
            leisu = build_leisu_crawler(headless=True)
            if leisu:
                inject_leisu_data(_target_match, leisu)
        except Exception as e:
            import traceback
            st.error(f"雷速数据注入失败: {e}")
            print(traceback.format_exc())

        if _action_type == 'repredict':
            with st.spinner("正在调用大模型重新进行全场预测..."):
                from src.llm.predictor import LLMPredictor
                predictor = LLMPredictor()
                total_count = len(filtered_matches)
                other_matches_context = [om for om in filtered_matches if om.get('match_num') != _target_match.get('match_num')]
                new_pred, period = predictor.predict(_target_match, total_matches_count=total_count, other_matches_context=other_matches_context)
                _target_match["llm_prediction"] = new_pred
                db = Database()
                db.save_prediction(_target_match, period)
                if "all_predictions" not in _target_match:
                    _target_match["all_predictions"] = {}
                _target_match["all_predictions"][period] = new_pred
                db.close()
                load_data.clear()
                save_data(matches)
            st.success("全场预测已更新！")
        elif _action_type == 'htft':
            with st.spinner("正在调用大模型进行专项半全场预测..."):
                from src.llm.htft_predictor import HTFTPredictor
                htft_predictor = HTFTPredictor()
                htft_pred, _ = htft_predictor.predict(_target_match, total_matches_count=1)
                _target_match["htft_prediction"] = htft_pred
                load_data.clear()
                save_data(matches)
            st.success("半全场预测已生成！")
        elif _action_type == 'goals':
            with st.spinner("正在调用大模型进行专项进球数推演..."):
                from src.llm.goals_predictor import GoalsPredictor
                goals_predictor = GoalsPredictor()
                try:
                    import pandas as pandas_lib
                    excel_path = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "docs", "foot_prediction.xlsx")
                    if os.path.exists(excel_path):
                        xl = pandas_lib.ExcelFile(excel_path)
                        for sheet_name in reversed(xl.sheet_names):
                            df_excel = pandas_lib.read_excel(xl, sheet_name=sheet_name)
                            df_excel.columns = df_excel.columns.str.strip()
                            if '编码' not in df_excel.columns:
                                continue
                            df_excel['编码'] = df_excel['编码'].astype(str).str.strip()
                            row = df_excel[df_excel['编码'] == _target_match.get('match_num')]
                            if not row.empty:
                                pan = row.iloc[0].get('进球盘口', '')
                                diff = row.iloc[0].get('预测差异百分比', '')
                                trend = row.iloc[0].get('倾向', '')
                                _target_match['goals_pan'] = str(pan).strip() if pandas_lib.notna(pan) else ""
                                _target_match['goals_diff_percent'] = str(diff).strip() if pandas_lib.notna(diff) else ""
                                _target_match['goals_trend'] = str(trend).strip() if pandas_lib.notna(trend) else ""
                                _target_match['excel_sheet_name'] = sheet_name
                                break
                except Exception as e:
                    st.warning(f"读取 Excel 数据失败: {e}")
                goals_pred, _ = goals_predictor.predict(_target_match)
                try:
                    res_dict = json.loads(goals_pred)
                    fund_report = res_dict.get('fundamental_report', goals_pred)
                except Exception:
                    fund_report = goals_pred
                _target_match["goals_prediction_special"] = fund_report
                load_data.clear()
                save_data(matches)
            st.success("进球数专项预测完成！")
        if leisu:
            try: leisu.close()
            except Exception: pass
        st.rerun()

    for match in filtered_matches:
        # 在这里也需要获取一次 recommendation
        recommendation = "待分析"
        pred_color = "gray"

        prediction_text, merged_all_preds = _resolve_primary_prediction_text(match)
        if merged_all_preds:
            match["all_predictions"] = merged_all_preds
            
        if prediction_text:
            from src.llm.predictor import LLMPredictor
            details = LLMPredictor.parse_prediction_details(prediction_text)
            rec_nspf = details.get('recommendation_nspf', '')
            if rec_nspf and rec_nspf != '暂无':
                recommendation = rec_nspf
                pred_color = "green" if "胜" in recommendation else ("blue" if "平" in recommendation else "red")
            else:
                recommendation = "已分析 (需点开查看)"
                pred_color = "black"
                
        with st.expander(f"📌 {match.get('match_num')} | {match.get('league')} | {match.get('home_team')} VS {match.get('away_team')} | {match.get('match_time')} | 预测: :{pred_color}[{recommendation}]", expanded=st.session_state.expand_all):
            
            col1, col2 = st.columns([1, 1])
            
            with col1:
                st.subheader("📊 基本面数据")
                recent = match.get("recent_form", {})
                st.markdown(f"**主队近期战绩:** {recent.get('home', '暂无数据')}")
                st.markdown(f"**客队近期战绩:** {recent.get('away', '暂无数据')}")
                st.markdown(f"**交锋记录:** {match.get('h2h_summary', '暂无数据')}")
                
                st.subheader("🎫 竞彩官方赔率")
                odds = match.get("odds", {})
                
                # 竞彩赔率表格
                jc_data = []
                if odds.get("nspf"):
                    jc_data.append(["不让球(0)"] + odds.get("nspf"))
                if odds.get("spf"):
                    jc_data.append([f"让球({odds.get('rangqiu')})"] + odds.get("spf"))
                    
                if jc_data:
                    df_jc = pd.DataFrame(jc_data, columns=["玩法", "胜", "平", "负"])
                    st.dataframe(df_jc, hide_index=True)
                else:
                    st.write("暂无竞彩赔率")

            with col2:
                st.subheader("📈 亚指盘口异动")
                asian = match.get("asian_odds", {})
                if not asian:
                    st.write("暂无亚指数据")
                else:
                    for company, data in asian.items():
                        comp_name = "澳门" if company == "macau" else "Bet365"
                        st.markdown(f"**{comp_name}**")
                        st.text(f"初盘: {data.get('start')}")
                        st.text(f"即时: {data.get('live')}")
                        st.divider()

            st.subheader("🤖 泊松深度预测报告")
            
            # 使用 Tabs 分离 全场预测 和 半全场专项预测
            pred_tab1, pred_tab2 = st.tabs(["🎯 全场赛果预测 (胜平负)", "🌗 半全场专项预测 (平胜/平负)"])
            
            with pred_tab1:
                # 获取不同时间段的预测结果
                db = Database()
                fixture_id = match.get("fixture_id")
            
                # 获取所有时间段的预测
                db_all_predictions = {}
                
                if fixture_id:
                    predictions = db.get_all_predictions_by_fixture(fixture_id)
                    for pred in predictions:
                        db_all_predictions[pred.prediction_period] = pred.prediction_text

                _, all_predictions = _resolve_primary_prediction_text(match, db_predictions=db_all_predictions)
                if all_predictions:
                    match["all_predictions"] = all_predictions
                
                # 显示时间段选择器
                prediction = ""
                if all_predictions:
                    periods = list(all_predictions.keys())
                    selected_period = st.selectbox("选择预测时间段", periods, 
                                                   format_func=lambda x: {
                                                       "pre_24h": "赛前24小时",
                                                       "pre_12h": "赛前12小时", 
                                                       "final": "赛前最终"
                                                   }.get(x, x),
                                                   key=f"period_select_{match.get('match_num')}")
                    
                    prediction = all_predictions.get(selected_period, "")
                    
                    # 如果有多个时间段，显示对比按钮
                    if len(all_predictions) > 1:
                        if st.button("🔍 对比分析不同时间段预测", key=f"compare_{match.get('match_num')}"):
                            _show_prediction_comparison(match, all_predictions)
                else:
                    prediction = match.get("llm_prediction", "")
            
                # 全场与半场逻辑一致性预警（红灯机制）
                htft_text = match.get("htft_prediction", "")
                if prediction and htft_text:
                    from src.llm.predictor import LLMPredictor
                    details = LLMPredictor.parse_prediction_details(prediction)
                    full_rec = details.get('recommendation', '')
                    
                    import re
                    htft_rec = ""
                    rec_match = re.search(r'半全场单关推荐[^\n]*?[:：]\s*\[?(.*?)\]?(?=\n|$)', htft_text)
                    if rec_match:
                        htft_rec = rec_match.group(1).replace('[', '').replace(']', '').replace('**', '').strip()
                    
                    if '放弃' not in htft_rec and htft_rec != "无":
                        conflict = False
                        
                        # 提取全场预测中的核心选项（去除权重和附加文字，提取真正的“胜/平/负”）
                        full_opts = set()
                        if '胜' in full_rec and '让胜' not in full_rec: full_opts.add('胜')
                        if '平' in full_rec and '让平' not in full_rec: full_opts.add('平')
                        if '负' in full_rec and '让负' not in full_rec: full_opts.add('负')
                        
                        # 半全场推荐打出的最终赛果必定是它的第二个字
                        htft_final_results = set()
                        if '胜' in htft_rec[-1:]: htft_final_results.add('胜')
                        if '平' in htft_rec[-1:]: htft_final_results.add('平')
                        if '负' in htft_rec[-1:]: htft_final_results.add('负')
                        
                        # 如果半全场推演的最终赛果，根本不在全场模型给出的双选/单选集合中，则判定为冲突
                        if htft_final_results and not (htft_final_results & full_opts):
                            conflict = True
                            
                        if conflict:
                            st.error(f"🚨 **模型分歧预警**：全场模型倾向于【{full_rec}】，而半全场专项模型倾向于【{htft_rec}】。两套独立推演逻辑出现严重冲突，说明本场比赛极具迷惑性或存在巨大变数，**建议直接放弃该场比赛的投注！**")

                if prediction:
                    if "Authentication Fails" in prediction or "预测失败" in prediction:
                        st.error(prediction)
                        st.info("提示: 请在 config/.env 文件中配置有效的 LLM_API_KEY。")
                    else:
                        st.markdown(prediction)
                else:
                    st.info("该场比赛尚未生成 AI 预测结果。")
            
            with pred_tab2:
                htft_prediction = match.get("htft_prediction", "")
                if htft_prediction:
                    st.markdown(htft_prediction)
                else:
                    st.info("该场比赛尚未生成半全场专项预测，请尝试重新执行全局预测。")
                
            # 显示独立进球数预测结果
            goals_prediction_special = match.get("goals_prediction_special", "")
            if goals_prediction_special:
                st.divider()
                st.markdown("### ⚽ 专项进球数推演报告 (来自 Excel 后验数据)")
                st.markdown(goals_prediction_special)
                
            # 单场重新预测功能 (仅 Admin)
            if st.session_state.get("role") == "admin":
                st.divider()
                col_rp1, col_rp2, col_rp3 = st.columns([1, 1, 1])
                with col_rp1:
                    if st.button("🔄 重新预测此场比赛 (全场)", key=f"repredict_{match.get('match_num')}"):
                        st.session_state[f"repredict_pending_{match.get('match_num')}"] = True
                        st.rerun()

                with col_rp2:
                    if st.button("🌗 触发专项半全场预测", key=f"repredict_htft_{match.get('match_num')}"):
                        st.session_state[f"repredict_htft_pending_{match.get('match_num')}"] = True
                        st.rerun()

                with col_rp3:
                    if st.button("⚽ 触发专项进球数推演", key=f"repredict_goals_{match.get('match_num')}"):
                        st.session_state[f"repredict_goals_pending_{match.get('match_num')}"] = True
                        st.rerun()

                # --- 旧版执行块已移至页面顶层，此处不再执行 ---

            article_key = f"article_content_{match.get('match_num')}"
            if article_key not in st.session_state:
                st.session_state[article_key] = match.get("article_text", "")
                
            saved_article = st.session_state[article_key]
            
            # 处理因为之前历史数据遗留导致的列表格式问题
            if isinstance(saved_article, list):
                saved_article = saved_article[0] if len(saved_article) > 0 else ""
                st.session_state[article_key] = saved_article
                
            if saved_article and st.session_state.get("role") == "admin":
                st.divider()
                st.markdown("### 📝 已生成的公众号推文")
                
                tab1, tab2 = st.tabs(["📋 Markdown 源码 (推荐使用排版工具)", "👁️ 网页预览 (可尝试直接复制)"])
                
                with tab1:
                    st.info("💡 **微信公众号排版指南**：\n\n微信编辑器不支持直接粘贴带有 `#` 或 `**` 的 Markdown 源码。业界标准做法是：\n1. 复制下方代码\n2. 打开微信排版神器，如 **[Markdown Nice (mdnice.com)](https://mdnice.com/)** 或 **[Doocs (doocs.github.io/md)](https://doocs.github.io/md/)**\n3. 粘贴进去，选择一个体育风格的主题，一键同步或复制到微信后台，效果极其精美！")
                    
                    # 重新使用 text_area，因为 st.code 在某些 Streamlit 版本下如果是长文本且没有换行时会显示成横向滚动的一长行
                    # 关键修复：让 text_area 绑定到一个专用的 key，以避免和外部的数据冲突
                    text_area_key = f"textarea_{match.get('match_num')}"
                    if text_area_key not in st.session_state or st.session_state.get(f"update_flag_{match.get('match_num')}"):
                        st.session_state[text_area_key] = saved_article
                        st.session_state[f"update_flag_{match.get('match_num')}"] = False
                        
                    st.text_area("Markdown 源码：", key=text_area_key, height=500)
                
                with tab2:
                    st.success("💡 提示：如果您不想使用第三方排版工具，可以尝试直接用鼠标从下方**选中所有文字**进行复制，然后直接粘贴到微信公众号后台（可以保留基础的加粗和标题排版样式）。")
                    
                    # 先将文本中的单个换行符替换为 Markdown 原生的强制换行语法（即在行尾加两个空格然后换行），
                    # 将双换行符（段落）替换为 HTML 的段落标签或保留，确保带 HTML 标签的混合渲染不会吞掉换行
                    # 因为开启了 unsafe_allow_html=True 后，部分 Markdown 解析器会进入 HTML 模式从而忽略 \n
                    
                    # 替换处理：先处理段落，再处理单行换行
                    html_article = saved_article.replace('\n\n', '<br><br>').replace('\n', '<br>')
                    
                    # 使用基础的 st.markdown 进行渲染，并且开启 unsafe_allow_html=True 使得标红代码生效
                    st.markdown(f"<div style='line-height: 1.8;'>{html_article}</div>", unsafe_allow_html=True)
            
            # 只有 admin 才能看到“生成/重新生成公众号文章”按钮
            if st.session_state.get("role") == "admin" and prediction:
                st.divider()
                btn_text = "🔄 重新生成公众号推文" if saved_article else "📝 一键生成公众号推文"
                if st.button(btn_text, key=f"btn_{match.get('match_num')}"):
                    with st.spinner("正在生成专业前瞻推文..."):
                        import sys
                        import os
                        from src.llm.predictor import LLMPredictor
                        predictor = LLMPredictor()
                        article, _ = predictor.generate_article(match, prediction, all_matches=matches)
                        
                        # 保存到内存字典
                        match["article_text"] = article
                        
                        # 清除缓存，强制重新读取最新的 JSON 数据
                        load_data.clear()
                        save_data(matches)
                        
                        # 保存到数据库
                        db = Database()
                        from src.llm.predictor import LLMPredictor
                        current_period = LLMPredictor()._determine_prediction_period(match)
                        db.save_prediction(match, current_period)
                        db.close()
                        
                        # 存入 Session State，必须在 rerun 之前！
                        st.session_state[article_key] = article
                        # 设置一个标志，强制刷新 text_area 的值
                        st.session_state[f"update_flag_{match.get('match_num')}"] = True
                        
                        st.success("生成成功并已持久化保存！")
                        st.rerun()

def _show_prediction_comparison(match, all_predictions):
    """显示不同时间段的预测对比分析"""
    import streamlit as st
    import pandas as pd
    from src.llm.predictor import LLMPredictor
    st.markdown("### 🔍 预测对比分析")
    
    # 按时间段排序
    periods = ['pre_24h', 'pre_12h', 'final']
    period_names = {
        'pre_24h': '赛前24小时',
        'pre_12h': '赛前12小时', 
        'final': '赛前最终'
    }
    
    # 创建对比表格
    comparison_data = []
    for period in periods:
        if period in all_predictions:
            pred_text = all_predictions[period]
            # 提取关键信息
            details = LLMPredictor.parse_prediction_details(pred_text)
            
            comparison_data.append({
                '时间段': period_names.get(period, period),
                '竞彩推荐': details.get('recommendation', '无'),
                '进球数参考': details.get('goals', '无'),
                '比分参考': details.get('score', '无'),
                '信心指数': details.get('confidence', '无'),
                '基础理由': details.get('reason', '无')
            })
    
    if comparison_data:
        df_comparison = pd.DataFrame(comparison_data)
        st.dataframe(df_comparison, use_container_width=True)
        
        # 分析变化
        if len(comparison_data) >= 2:
            st.markdown("#### 📊 变化分析")
            changes = _analyze_prediction_changes(comparison_data)
            st.markdown(changes)
    else:
        st.info("暂无足够的数据进行对比分析")

def _analyze_prediction_changes(comparison_data):
    """分析预测变化并提供原因"""
    analysis = ""
    
    # 检查预测结果是否发生变化
    results = [data['竞彩推荐'] for data in comparison_data]
    if len(set(results)) > 1:
        analysis += "🔀 **推荐方向发生变化**：\n"
        analysis += f"- 从 【{results[0]}】 变为 【{results[-1]}】\n"
        analysis += "- 可能原因：盘口剧烈变动、临场首发名单确认或资金极度倾斜导致机构调整预期。\n\n"
    
    # 检查信心指数变化
    confidences = [data['信心指数'] for data in comparison_data]
    if len(set(confidences)) > 1:
        analysis += "📈 **信心指数变化**：\n"
        analysis += f"- 从 {confidences[0]} 变为 {confidences[-1]}\n"
        # 简单比对长度（星星数量）
        if len(confidences[-1]) > len(confidences[0]):
            analysis += "- 机构数据趋于明朗，预测确定性增强，冷门概率降低。\n"
        elif len(confidences[-1]) < len(confidences[0]):
            analysis += "- 出现新的不确定因素或盘口走势与基本面背离，需要谨慎对待防范冷门。\n\n"
    
    if not analysis:
        analysis = "✅ **预测保持高度一致**：各时间段的核心推荐和信心指数均保持稳定，说明比赛基本面和盘口逻辑没有发生根本性反转，可以按原计划参考。"
    
    return analysis

if __name__ == "__main__":
    main()
