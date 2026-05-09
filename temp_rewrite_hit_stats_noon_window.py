import sqlite3
from datetime import datetime, timedelta
from collections import defaultdict
from openpyxl import load_workbook

DB_PATH = r'e:\zhangxuejun\football-prediction\data\football.db'
XLSX_PATH = r'e:\zhangxuejun\football-prediction\docs\泊松分析系统命中率统计.xlsx'

conn = sqlite3.connect(DB_PATH)
cur = conn.cursor()
cur.execute('''
    SELECT id, fixture_id, match_num, home_team, away_team, match_time, actual_result, predicted_result, created_at
    FROM match_predictions
    WHERE predicted_result IS NOT NULL
      AND TRIM(predicted_result) != ''
''')
rows = cur.fetchall()
conn.close()

def parse_dt(text):
    if not text:
        return None
    text = str(text).strip()
    for fmt in ('%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M'):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    return None

latest = {}
for row in rows:
    row_id, fixture_id, match_num, home_team, away_team, match_time, actual_result, predicted_result, created_at = row
    mt = parse_dt(match_time)
    match_key = fixture_id or f'{match_num}|{home_team}|{away_team}|{str(match_time)[:10]}'
    current = latest.get(match_key)
    row_dt = parse_dt(created_at) or datetime.min
    if current is None or row_dt > current['created_at'] or (row_dt == current['created_at'] and row_id > current['id']):
        latest[match_key] = {
            'id': row_id,
            'created_at': row_dt,
            'match_time': mt,
            'actual_result': (actual_result or '').strip(),
            'predicted_result': (predicted_result or '').strip(),
        }

stats = defaultdict(lambda: {'total': 0, 'hit': 0})
for item in latest.values():
    mt = item['match_time']
    actual = item['actual_result']
    pred = item['predicted_result']
    if not mt or actual not in ('胜', '平', '负'):
        continue
    stat_date = (mt - timedelta(hours=12)).date().isoformat()
    stats[stat_date]['total'] += 1
    if actual in pred:
        stats[stat_date]['hit'] += 1

ordered = sorted(stats.items())

wb = load_workbook(XLSX_PATH)
ws = wb['Sheet1']
if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row - 1)

for idx, (stat_date, value) in enumerate(ordered, start=2):
    total = value['total']
    hit = value['hit']
    hit_rate = hit / total if total else 0
    ws.cell(row=idx, column=1, value=stat_date)
    ws.cell(row=idx, column=2, value=total)
    ws.cell(row=idx, column=3, value=hit)
    ws.cell(row=idx, column=4, value=f'{hit_rate:.2%}')

wb.save(XLSX_PATH)

print('MATCHES', len(latest))
print('DAYS', len(ordered))
print('FIRST5', ordered[:5])
print('LAST5', ordered[-5:])
