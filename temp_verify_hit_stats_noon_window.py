import openpyxl
path = r'e:\zhangxuejun\football-prediction\docs\泊松分析系统命中率统计.xlsx'
wb = openpyxl.load_workbook(path)
ws = wb['Sheet1']
print('ROWS', ws.max_row, ws.max_column)
for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row,14), values_only=True):
    print(row)
